from write import *
import openpyxl
wb2 = openpyxl.load_workbook("Abdominal Pain Rules.xlsx")
sheet = wb2.active
row_counts = sheet.max_row
col_counts= sheet.max_column
#1 row iterate
dic={}
arr=[]

        
for i in range(1,col_counts+1):
    arry=sheet.cell(row=1, column=i).value
    arr.append(arry.lower())

def check(values,arrays):
    keywords=""
    count=0
    for i in range(len(arrays)):
        for j in range(len(values)-2):
            if arrays[i]==values[j]:
                keywords+=str(arrays[i])+", \n"
                count+=1
    return count,keywords
####Yhaan tk separate
def compare(arrays):
    keywordds=""
    countt=0
    dicc={}
    for i in range(2,row_counts+1):
        count=0
        dic={}
        for j in range(1,col_counts+1):
            x=sheet.cell(row=i, column=j).value
            if x==None:
                dic[str(arr[j-1])]="."
            else:
                dic[arr[j-1]]=x.lower()

            values=list(dic.values())
        count,keywords=check(values,arrays)
        if count>countt:
            countt=count
            dicc=dic
            keywordds=keywords

    #print(list(dicc.values()))
    #print(countt,keywordds)
    #print("Diagnosis: "+dicc["diagnosis"]+"\n"+"Disposition: "+dicc["disposition"])
    write(dicc["diagnosis"],dicc["disposition"],str(countt),keywordds)
#arrays=["none","none","none","none","none","none","none","none","none","none","none","none","none","diarrhea, >6 bm/day","none","none","non-tender","none"]
#compare(arrays)

def my_show():
    complaint1=complaint_opt.get().lower()
    PMHx2=PMHx_opt.get().lower()
    Medic3=Medic_opt.get().lower()
    allerg4=Allerg_opt.get().lower()
    surg5=Surg_opt.get().lower()
    social6=social_opt.get().lower()
    fam7=fam_opt.get().lower()
    pen8=pen_opt.get().lower()
    ons9=ons_opt.get().lower()
    age10=age_opt.get().lower()
    preg11=preg_opt.get().lower()
    brea12=brea_opt.get().lower()
    tim13=time_opt.get().lower()
    hpi14=hpi_opt.get().lower()
    ros15=ros_opt.get().lower()
    vitt16=vit_opt.get().lower()
    exam17=exam_opt.get().lower()
    proc18=proc_opt.get().lower()
    arrays=[complaint1,PMHx2,Medic3,allerg4,surg5,social6,fam7,pen8,ons9,age10,preg11,brea12,tim13,hpi14,ros15,vitt16,exam17,proc18]
    compare(arrays)
from tkinter import *
win = Tk()
win.geometry("700x600")  # Size of the window 
win.title("Medical App")  # Adding a title
win.resizable(width=False, height=False)
bg = PhotoImage(file = "background/back.png")
label1 = Label( win, image = bg)
label1.place(x = 0, y = 0)
####



#1st column
complaint_label = Label(win,  text='Complaint', width=15 )  
complaint_label.grid(row=1,column=1)

complaint_List = ["abdominal pain","None"]
complaint_opt = StringVar(win)
complaint_opt.set(complaint_List[0]) # default value
complaint_Menu =OptionMenu(win, complaint_opt, *complaint_List)
complaint_Menu.grid(row=1,column=2)

#2nd column
PMHx_label = Label(win,  text='Select PMHx', width=15 )  
PMHx_label.grid(row=2,column=1)

PMHx_List = ["cyst","None"]
PMHx_opt = StringVar(win)
PMHx_opt.set(PMHx_List[0]) # default value
PMHx_Menu =OptionMenu(win, PMHx_opt, *PMHx_List)
PMHx_Menu.grid(row=2,column=2)


#3rd column
Medic_label = Label(win,  text='Select Medication', width=15 )  
Medic_label.grid(row=3,column=1)

Medic_List = ["truvada","None"]
Medic_opt = StringVar(win)
Medic_opt.set(Medic_List[0]) # default value
Medic_Menu =OptionMenu(win, Medic_opt, *Medic_List)
Medic_Menu.grid(row=3,column=2)


#4th column
Allerg_label = Label(win,  text='Select Allergies', width=15 )  
Allerg_label.grid(row=4,column=1)

Allerg_List = ["None"]
Allerg_opt = StringVar(win)
Allerg_opt.set(Allerg_List[0]) # default value
Allerg_Menu =OptionMenu(win, Allerg_opt, *Allerg_List)
Allerg_Menu.grid(row=4,column=2)


#5th column
Surg_label = Label(win,  text='Select Surgery', width=15 )  
Surg_label.grid(row=5,column=1)

Surg_List = ["bariatric surgery","None"]
Surg_opt = StringVar(win)
Surg_opt.set(Allerg_List[0]) # default value
Surg_Menu =OptionMenu(win, Surg_opt, *Surg_List)
Surg_Menu.grid(row=5,column=2)

#6th column
social_label = Label(win,  text='Select Social Hx', width=15 )  
social_label.grid(row=6,column=1)

social_List = ["None"]
social_opt = StringVar(win)
social_opt.set(social_List[0]) # default value
social_Menu =OptionMenu(win, social_opt, *social_List)
social_Menu.grid(row=6,column=2)


#7th column
fam_label = Label(win,  text='Select Family Hx', width=15 )  
fam_label.grid(row=7,column=1)

fam_List = ["None"]
fam_opt = StringVar(win)
fam_opt.set(fam_List[0]) # default value
fam_Menu =OptionMenu(win, fam_opt, *fam_List)
fam_Menu.grid(row=7,column=2)

#8th column
pen_label = Label(win,  text='Select pen Allergy', width=15 )  
pen_label.grid(row=8,column=1)

pen_List = ["None"]
pen_opt = StringVar(win)
pen_opt.set(pen_List[0]) # default value
pen_Menu =OptionMenu(win, pen_opt, *pen_List)
pen_Menu.grid(row=8,column=2)



#9th column
ons_label = Label(win,  text='Select Onset', width=15 )  
ons_label.grid(row=9,column=1)

ons_List = ["None"]
ons_opt = StringVar(win)
ons_opt.set(ons_List[0]) # default value
ons_Menu =OptionMenu(win, ons_opt, *ons_List)
ons_Menu.grid(row=9,column=2)

####Half way done

#10th column
age_label = Label(win,  text='Select Age', width=15 )  
age_label.grid(row=1,column=3)

age_List = ["age: <60","age: >50","age: <19","age: >10","None"]
age_opt = StringVar(win)
age_opt.set(age_List[0]) # default value
age_Menu =OptionMenu(win, age_opt, *age_List)
age_Menu.grid(row=1,column=4)

#11th column
preg_label = Label(win,  text='Pregnancy', width=15 )  
preg_label.grid(row=2,column=3)

preg_List = ["pregnant","female","None"]
preg_opt = StringVar(win)
preg_opt.set(preg_List[0]) # default value
preg_Menu =OptionMenu(win, preg_opt, *preg_List)
preg_Menu.grid(row=2,column=4)

#12th column
brea_label = Label(win,  text='Breast', width=15 )  
brea_label.grid(row=3,column=3)

brea_List = ["None"]
brea_opt = StringVar(win)
brea_opt.set(brea_List[0]) # default value
brea_Menu =OptionMenu(win, brea_opt, *brea_List)
brea_Menu.grid(row=3,column=4)

#13th column
time_label = Label(win,  text='TimeOfYear', width=15 )  
time_label.grid(row=4,column=3)

time_List = ["None"]
time_opt = StringVar(win)
time_opt.set(time_List[0]) # default value
time_Menu =OptionMenu(win, time_opt, *time_List)
time_Menu.grid(row=4,column=4)

#14th column
hpi_label = Label(win,  text='Select HPI_1', width=15 )  
hpi_label.grid(row=5,column=3)

hpi_List = ["(+) abdominal pain","ill appearing","10/10","abdominal pain",
            "(+) nausea","vaginal bleed","flank pain","female","male",
            "abdominal surgery","pain and worse or better with  food",
            "diarrhea, >6 bm/day","diarrhea and travel","diarrhea, > 1 week",
            "diarrhea and giardia exposure","unable to get comfortable",
            "nausea","constant pain","flank pain","None"]
hpi_opt = StringVar(win)
hpi_opt.set(hpi_List[0]) # default value
hpi_Menu =OptionMenu(win, hpi_opt, *hpi_List)
hpi_Menu.grid(row=5,column=4)

#15th column
ros_label = Label(win,  text='Select ROS', width=15 )  
ros_label.grid(row=6,column=3)

ros_List = ["None"]
ros_opt = StringVar(win)
ros_opt.set(ros_List[0]) # default value
ros_Menu =OptionMenu(win, ros_opt, *ros_List)
ros_Menu.grid(row=6,column=4)

#16th column
vit_label = Label(win,  text='Select vital', width=15 )  
vit_label.grid(row=7,column=3)

vit_List = ["hr:more120","bp:less90/60","temp:more than 102.0","None"]
vit_opt = StringVar(win)
vit_opt.set(vit_List[0]) # default value
vit_Menu =OptionMenu(win, vit_opt, *vit_List)
vit_Menu.grid(row=7,column=4)

#17th column
exam_label = Label(win,  text='Select Exam', width=15 )  
exam_label.grid(row=8,column=3)

exam_List = ["diffuse tenderness, no rebound/guarding","diffuse tenderness and (rebound or guarding)",
             "diffusely tender","diffusely tender and (rebound or guarding)","non-tender","LUQ or epigastric or periumbilical tenderness, no rebound or guarding",
             "abdominal tenderness, Adnexal Tenderness","pelvic exam normal","tenderness","Left lower quadrant tenderness, CMT, adnexal tenderness",
             "RUQ tenderness","jaundice","RLQ tenderness","adnexal tenderness","Tenderness, no rebound/guarding","testicle pain","pulse deficit",
             "neuro deficit","LLQ tenderness","tenderness","None"]
exam_opt = StringVar(win)
exam_opt.set(exam_List[0]) # default value
exam_Menu =OptionMenu(win, exam_opt, *exam_List)
exam_Menu.grid(row=8,column=4)

#18th column
proc_label = Label(win,  text='Procedure value', width=15 )  
proc_label.grid(row=9,column=3)

proc_List = ["ucg negative","ucg positive","ua no leukocytes","HIV positive","diverticulitis","leukocytes","None"]
proc_opt = StringVar(win)
proc_opt.set(proc_List[0]) # default value
proc_Menu =OptionMenu(win, proc_opt, *proc_List)
proc_Menu.grid(row=9,column=4)


##Run here
Button(win,text="START",command=my_show,bg="white",height = 3, width = 8).grid()
mainloop()























